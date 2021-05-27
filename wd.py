
import openpyxl
# import datetime
from datetime import datetime
from datetime import timedelta
import calendar

ln = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
# calendar.setfirstweekday(calendar.SUNDAY)
# print(calendar.weekday(2020, 7, 17))
# # 获取当前时间
# today = datetime.now()
# print(today)

# # 计算偏移量
# offset = timedelta(days=-2)
# print(offset)
#
# # 获取想要的日期的时间
# re_date = (today + offset).strftime('%Y-%m-%d')
# print(re_date)

# time1= "2019-5-12 12:13:14"		# 字符串 日期
# d1 = datetime.strptime(str(time1),'%Y-%m-%d %H:%M:%S')
# print(type(d1))
# plus= d1 + timedelta(days=1)		# 加
# minus = d1 - timedelta(days=1)		# 减
# print(time1)
# print(d1)
# print(plus)
# print(minus )


def getDate(date):
    y_m_d = str(date).split(' ')
    y_m_d = y_m_d[0].replace('-', '/')

    return y_m_d

# '''
if __name__ == "__main__":
    try:
        wb = openpyxl.load_workbook('汪聪2021周志.xlsx')
    except:
        print('无法加载文件')
    ws = wb.active
    latest_row = ws.max_row
    # print(latest_row)

    start_row = latest_row + 2
    ld = ws.cell(row=latest_row, column=1).value

    # print(ld)
    # a = str(time1).split(' ')
    latest_date = datetime.strptime(ld, '%Y/%m/%d')
    # print(type(latest_date))
    # print(latest_date)
    # plus= latest_date + timedelta(days=1)		# 加


    for i in range(6):
        ws.cell(row=start_row + i, column=2).value = ln[i]
        # print(ln[i])
        ws.cell(row=start_row + i, column=1).value = getDate(latest_date + timedelta(days=i+2))
        # print(getDate(latest_date + timedelta(days=i+2)))



    wb.save(filename = '汪聪2021周志.xlsx')
    print('Everything is OK!')
# '''


